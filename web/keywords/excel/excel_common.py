from openpyxl import load_workbook
import os, sys
web_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../"))
sys.path.append(web_path)

class Excel:
    @staticmethod
    def get_file_path(file_name):
        return os.path.join(web_path, "resources", "testdata", file_name)
    
    @staticmethod
    def get_login_data():
        file_path = Excel.get_file_path("test_data.xlsx")
        workbook = load_workbook(file_path)
        sheet = workbook["LoginData"]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            email, password = row[:2]
            data.append((email, password))
        print(data)
        return data

    @staticmethod
    def write_data_to_excel(email, password, result):
        file_path = Excel.get_file_path("test_data.xlsx")
        workbook = load_workbook(file_path)
        sheet = workbook["SignUpData"]
        sheet.append([email, password, result])
        workbook.save(file_path)
